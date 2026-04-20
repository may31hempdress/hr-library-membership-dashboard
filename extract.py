"""
週次/月次メンバー数データを Excel から抽出し data.json に出力する。

使い方:
    python extract.py <xlsx_path>
    # デフォルト: ../参考資料/数字で見る人事図書館変遷_*.xlsx の最新ファイル

出力:
    data.json
"""
import sys
import json
import glob
import os
from datetime import datetime, date, time
import openpyxl


def to_iso(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def clean_num(v):
    if v is None or v == "" or v == "#N/A":
        return None
    if isinstance(v, time):
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None


def extract_series(ws, row_idx, start_col, end_col):
    row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
    return [clean_num(row[c]) for c in range(start_col, end_col + 1)]


def extract_dates(ws, row_idx, start_col, end_col):
    row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
    return [to_iso(row[c]) for c in range(start_col, end_col + 1)]


def find_row(ws, label, search_range=(1, 40)):
    for ri, row in enumerate(
        ws.iter_rows(
            min_row=search_range[0], max_row=search_range[1], max_col=1, values_only=True
        ),
        search_range[0],
    ):
        cell = row[0]
        if cell and label in str(cell):
            return ri
    return None


def extract_sheet(ws, periodicity):
    # Rows are identified by the label in col 0; dates live in the first row
    # with date values.
    # Find date row (first row whose col 0 is empty but col 1 is a date)
    date_row_idx = None
    for ri in range(1, 5):
        row = next(ws.iter_rows(min_row=ri, max_row=ri, max_col=3, values_only=True))
        if (row[0] is None or row[0] == "") and isinstance(row[1], datetime):
            date_row_idx = ri
            break
    if date_row_idx is None:
        raise RuntimeError(f"date row not found in {periodicity}")

    # Determine column range for dates
    date_row = next(
        ws.iter_rows(min_row=date_row_idx, max_row=date_row_idx, values_only=True)
    )
    last_col = 0
    for i, c in enumerate(date_row):
        if isinstance(c, datetime):
            last_col = i
    dates = extract_dates(ws, date_row_idx, 1, last_col)

    # Rows we need — labels as they appear in the sheet (partial match)
    wanted = {
        "cum_total": "①のべメンバー数",  # (cumulative total)
        "cum_full": "①ーA",
        "cum_online": "①ーB",
        "cum_corp": "①ーC",
        "cum_leave_full": "②ーA",
        "cum_leave_online": "②ーB",
        "cum_leave_corp": "②ーC",
        "active_total": "③アクティブメンバー数",
        "active_full": "③ーA",
        "active_online": "③ーB",
        "active_corp": "③ーC",
    }

    # ①のべ appears twice (cumulative and 増加分). We want the cumulative one,
    # which is the FIRST occurrence before any "増加分" row.
    # ②ーA etc. also appear twice (cumulative and 増加分 delta).
    # Strategy: scan sequentially, tracking whether we're in a "増加分" block.
    rows_info = []
    in_delta_block = False
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=50, max_col=2, values_only=True), 1):
        label = row[0]
        if label is None:
            continue
        s = str(label)
        if "増加分" in s:
            in_delta_block = True
            continue
        if "②総退会依頼数" in s or "現メンバー数" in s or "①" == s[:1] and s == "①のべメンバー数" and not in_delta_block:
            pass
        if s == "②総退会依頼数":
            in_delta_block = False
            continue
        if s == "現メンバー数":
            in_delta_block = False
            continue
        rows_info.append((ri, s, in_delta_block))

    def find(label_substr, delta):
        for ri, s, is_delta in rows_info:
            if label_substr in s and is_delta == delta:
                return ri
        return None

    out = {"dates": dates}

    # Cumulative 入会 / アクティブ (non-delta)
    for key, label in [
        ("cum_full", "①ーA"),
        ("cum_online", "①ーB"),
        ("cum_corp", "①ーC"),
        ("cum_leave_full", "②ーA"),
        ("cum_leave_online", "②ーB"),
        ("cum_leave_corp", "②ーC"),
        ("active_full", "③ーA"),
        ("active_online", "③ーB"),
        ("active_corp", "③ーC"),
    ]:
        ri = find(label, False)
        if ri is None:
            out[key] = [None] * len(dates)
            continue
        out[key] = extract_series(ws, ri, 1, last_col)

    # 増加分 (per-period) 入会 / 退会
    for key, label in [
        ("new_full", "①ーA"),
        ("new_online", "①ーB"),
        ("new_corp", "①ーC"),
        ("leave_full", "②ーA"),
        ("leave_online", "②ーB"),
        ("leave_corp", "②ーC"),
    ]:
        ri = find(label, True)
        if ri is None:
            out[key] = [None] * len(dates)
            continue
        out[key] = extract_series(ws, ri, 1, last_col)

    return out


def main():
    if len(sys.argv) > 1:
        xlsx_path = sys.argv[1]
    else:
        pattern = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "..",
            "参考資料",
            "数字で見る人事図書館変遷_*.xlsx",
        )
        candidates = sorted(glob.glob(pattern))
        if not candidates:
            print(f"no file matched: {pattern}", file=sys.stderr)
            sys.exit(1)
        xlsx_path = candidates[-1]

    print(f"reading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    weekly = extract_sheet(wb["週次グラフ用"], "weekly")
    monthly = extract_sheet(wb["月次グラフ用"], "monthly")

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "public", "data.json")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source": os.path.basename(xlsx_path),
        "weekly": weekly,
        "monthly": monthly,
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"wrote: {out_path}")
    print(f"  weekly periods: {len(weekly['dates'])} ({weekly['dates'][0]} .. {weekly['dates'][-1]})")
    print(f"  monthly periods: {len(monthly['dates'])} ({monthly['dates'][0]} .. {monthly['dates'][-1]})")


if __name__ == "__main__":
    main()
